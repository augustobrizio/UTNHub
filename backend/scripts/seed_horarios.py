"""Seed de comisiones y horarios desde los Excel de la FRRO.

Parsea los archivos Excel por año/comisión y carga las tablas
``comision``, ``cursada`` y ``horario`` en Neon.

Uso (local, conecta directo a Neon via DATABASE_URL en backend/.env):
    cd backend
    python scripts/seed_horarios.py [EXCEL_DIR] [--reset]

    EXCEL_DIR: directorio con los xlsx (default: C:/Users/Bruno/Downloads)
    --reset  : borra todas las comisiones y horarios antes de insertar

Uso (en Docker — primero copiar/montar los Excel):
    docker compose exec backend uv run python scripts/seed_horarios.py /tmp/horarios --reset
"""
from __future__ import annotations

import argparse
import datetime
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

# Permite importar app.* corriendo el script desde cualquier directorio.
sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env")

import openpyxl
from sqlalchemy import delete

from app.db.models.academico import Comision, Cursada, Horario
from app.db.session import SessionLocal


# ---------------------------------------------------------------------------
# Archivos Excel a procesar
# ---------------------------------------------------------------------------

ARCHIVOS: list[str] = [
    "Horarios 1º año (2).xlsx",
    "Horarios 2º año (1).xlsx",
    "Horarios 3º año (2).xlsx",
    "Horarios Analista 3º año (1).xlsx",
    "Horarios 4º año (1).xlsx",
    "Horarios 5º año (1).xlsx",
    "Horarios electivas 2º año (2).xlsx",
    "Horarios electivas 3º año (2).xlsx",
    "Horarios electivas 4º año (2).xlsx",
    "Horarios electivas 5º año (4).xlsx",
]

DIAS = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado"]
DIA_COL_START = 2  # col 0=Horario, 1=Horas, 2=Lunes ... 7=Sábado


# ---------------------------------------------------------------------------
# Normalización y mapeo de nombres a códigos de materia
# ---------------------------------------------------------------------------

def _n(s: str) -> str:
    """Minúsculas, sin acentos, sin espacios extra."""
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


NOMBRE_A_CODIGO: dict[str, str] = {
    # 1er año troncales
    _n("Análisis Matemático I"): "1",
    _n("Álgebra y Geometría Analítica"): "2",
    _n("Algebra y Geometría Analítica"): "2",
    _n("Algebra y  Geometría  Analítica"): "2",
    _n("Física I"): "3",
    _n("Inglés I"): "4",
    _n("Lógica y Estructuras Discretas"): "5",
    _n("Algoritmos y Estructuras de Datos"): "6",
    _n("Arquitectura de Computadoras"): "7",
    _n("Sistemas y Procesos de Negocio"): "8",
    # 2do año troncales
    _n("Análisis Matemático II"): "9",
    _n("Física II"): "10",
    _n("Ingeniería y Sociedad"): "11",
    _n("Inglés II"): "12",
    _n("Sintaxis y Semántica de los Lenguajes"): "13",
    _n("Paradigmas de Programación"): "14",
    _n("Sistemas Operativos"): "15",
    _n("Análisis de Sistemas de Información"): "16",
    # 3er año troncales
    _n("Probabilidad y Estadística"): "17",
    _n("Economía"): "18",
    _n("Base de Datos"): "19",
    _n("Bases de Datos"): "19",
    _n("Desarrollo de Software"): "20",
    _n("Comunicación de Datos"): "21",
    _n("Análisis Numérico"): "22",
    _n("Diseño de Sistemas de Información"): "23",
    _n("Diseño deSistemas de Información"): "23",
    _n("Diseño de Sistemas de Informacion"): "23",
    _n("Seminario Integrador"): "ADUSI",
    _n("Seminario Integrador (ADUSI)"): "ADUSI",
    # 4to año troncales
    _n("Legislación"): "24",
    _n("Ingeniería y Calidad de Software"): "25",
    _n("Redes de Datos"): "26",
    _n("Investigación Operativa"): "27",
    _n("Simulación"): "28",
    _n("Tecnologías para la Automatización"): "29",
    _n("Administración de Sistemas de Información"): "30",
    # 5to año troncales
    _n("Inteligencia Artificial"): "31",
    _n("Ciencia de Datos"): "32",
    _n("Sistemas de Gestión"): "33",
    _n("Gestión Gerencial"): "34",
    _n("Seguridad en los Sistemas de Información"): "35",
    _n("Proyecto Final"): "36",
    # Electivas
    _n("Entornos Gráficos"): "E01",
    _n("Análisis y Diseño de Datos e Información"): "E02",
    _n("Sistemas de Información Geográfica"): "E03",
    _n("Sist. de Información Geográfica"): "E03",
    _n("Formación de Emprendedores"): "E04",
    _n("Algoritmos Genéticos"): "E05",
    _n("Informática Jurídica"): "E06",
    _n("Lenguaje de Programación JAVA"): "E07",
    _n("Lenguaje de Programación Java"): "E07",
    _n("Tecnologías de Desarrollo de Software IDE"): "E08",
    _n("Tecnologías de Desarrollo deSoft IDE"): "E08",
    _n("Gestión Ingenieril"): "E09",
    _n("Introducción a la Práctica Profesional"): "E10",
    _n("Química Aplicada a la Informática"): "E11",
    _n("Infraestructura Tecnológica"): "E12",
    _n("Soporte a las Bases de Datos con Programación Visual"): "E13",
    _n("Metodología de la Investigación"): "E14",
    _n("Metodologías Ágiles en el Desarrollo de Software"): "E15",
    _n("Fabricación Aditiva"): "E16",
    _n("Dirección de Recursos Humanos"): "E17",
    _n("Informática en la Administración Pública"): "E18",
    _n("Sistemas de Información Integrados para la Industria"): "E19",
    # Abreviaciones usadas en los Excel
    _n("Administración de Sist. de Información"): "30",
    _n("Informática en la Adm. Pública"): "E18",
    _n("Metodologías Agiles en el Des. de Soft."): "E15",
    _n("Seguridad en los Sist. de Información"): "35",
    _n("Simulación Flamini-Torres"): "28",
    _n("Simulación Leale-Torres"): "28",
    _n("Soporte a la Gestión de datos con P.Visual"): "E13",
    _n("Sist. de Inf. Integrados para la Industria"): "E19",
}

# Strings que aparecen en las celdas del Excel pero no son materias
NOMBRES_IGNORAR: set[str] = {
    _n("Inscribirse en la Comisión"),
}


def mapear_codigo(nombre: str) -> str | None:
    key = _n(nombre)
    if key in NOMBRES_IGNORAR:
        return "__ignorar__"
    if key in NOMBRE_A_CODIGO:
        return NOMBRE_A_CODIGO[key]
    # Substring fallback para abreviaciones o ligeras variaciones
    if len(key) > 10:
        for k, v in NOMBRE_A_CODIGO.items():
            if len(k) > 10 and (k in key or key in k):
                return v
    return None


# ---------------------------------------------------------------------------
# Parsing del Excel
# ---------------------------------------------------------------------------

def _es_inicio_bloque(row: tuple) -> bool:
    """Row A de un bloque horario: (time, int|'PH'|'PPH', ...)."""
    return (
        isinstance(row[0], datetime.time)
        and (isinstance(row[1], int) or row[1] in ("PH", "PPH"))
    )


def _minutos_entre(t1: datetime.time, t2: datetime.time) -> int:
    d1 = datetime.timedelta(hours=t1.hour, minutes=t1.minute)
    d2 = datetime.timedelta(hours=t2.hour, minutes=t2.minute)
    return int((d2 - d1).total_seconds() / 60)


def _parse_seccion(rows: list[tuple]) -> list[dict]:
    """
    Extrae slots crudos de una sección cuatrimestral.

    Cada bloque en el Excel ocupa 3 filas:
      Row A: (hora_inicio, hora_num, mat_lun, mat_mar, mat_mie, mat_jue, mat_vie, mat_sab)
      Row B: ('a', None,  mat_cont, ...)      ← 2da línea del nombre
      Row C: (hora_fin, None, docente, ...)   ← fin + docentes

    Algunos nombres ocupan las 3 filas (p.ej. "Ingeniería y | Sociedad"):
    Row A="Ingeniería ", Row B="y", Row C col="Sociedad". En ese caso Row C
    tiene el nombre, NO el docente. Esto se resuelve en el mapeo posterior.
    """
    slots: list[dict] = []
    n = len(rows)
    i = 0
    while i < n:
        row_a = rows[i]
        if not _es_inicio_bloque(row_a):
            i += 1
            continue

        hora_inicio: datetime.time = row_a[0]

        # Row B (línea de continuación del nombre)
        row_b: tuple | None = None
        if i + 1 < n and rows[i + 1][0] == "a":
            row_b = rows[i + 1]

        # Row C (hora fin + docentes)
        rc_idx = (i + 2) if row_b is not None else (i + 1)
        row_c: tuple | None = None
        if rc_idx < n and isinstance(rows[rc_idx][0], datetime.time) and rows[rc_idx][1] is None:
            row_c = rows[rc_idx]

        hora_fin: datetime.time | None = row_c[0] if row_c else None

        for col_offset, dia in enumerate(DIAS):
            col = DIA_COL_START + col_offset
            name_a = str(row_a[col] or "").strip() if col < len(row_a) else ""
            name_b = str(row_b[col] or "").strip() if (row_b and col < len(row_b)) else ""
            # Row C puede tener docente o 3ra parte del nombre — se decide en mapeo
            name_c = str(row_c[col] or "").strip() if (row_c and col < len(row_c)) else ""

            nombre_parcial = f"{name_a} {name_b}".strip()
            if not nombre_parcial:
                continue

            slots.append({
                "hora_inicio": hora_inicio,
                "hora_fin": hora_fin,
                "dia": dia,
                "nombre_parcial": nombre_parcial,
                "nombre_3part": f"{nombre_parcial} {name_c}".strip() if name_c else nombre_parcial,
                "docente_raw": name_c or None,
            })

        i += 3 if row_b is not None else 2

    return slots


def _merge_blocks(slots: list[dict]) -> list[dict]:
    """
    Fusiona slots consecutivos de la misma materia en el mismo día.

    Considera "consecutivos" si el gap entre hora_fin y la siguiente
    hora_inicio es ≤ 20 min (cubre descansos de 15 min entre horas).
    """
    grupos: dict[tuple, list[dict]] = defaultdict(list)
    for s in slots:
        grupos[(s["dia"], s["nombre_parcial"])].append(s)

    merged: list[dict] = []
    for (dia, nombre), grupo in grupos.items():
        grupo.sort(key=lambda s: (s["hora_inicio"].hour, s["hora_inicio"].minute))
        cur = dict(grupo[0])
        docente = cur["docente_raw"]
        for s in grupo[1:]:
            if cur["hora_fin"] is None:
                break
            gap = _minutos_entre(cur["hora_fin"], s["hora_inicio"])
            if 0 <= gap <= 20:
                cur["hora_fin"] = s["hora_fin"]
                if not docente and s["docente_raw"]:
                    docente = s["docente_raw"]
            else:
                cur["docente_raw"] = docente
                merged.append(cur)
                cur = dict(s)
                docente = s["docente_raw"]
        cur["docente_raw"] = docente
        merged.append(cur)

    return merged


def parse_sheet(ws) -> list[dict]:
    """
    Parsea una sheet de comisión y devuelve lista de registros:
      {comision_nombre, cuatrimestre, nombre_parcial, nombre_3part,
       dia, hora_inicio, hora_fin, docente_raw}
    """
    all_rows = [row for row in ws.iter_rows(values_only=True) if any(c is not None for c in row)]

    # Metadata: comisión y turno
    comision_nombre: str | None = None
    for row in all_rows[:10]:
        if len(row) > 3 and str(row[2] or "").strip() == "Comisión" and row[3]:
            comision_nombre = str(row[3]).strip()
            break
    if not comision_nombre:
        return []

    # Localizar inicio de cada cuatrimestre
    secciones: list[tuple[int, int, int | None]] = []  # (num, start_idx, end_idx)
    for idx, row in enumerate(all_rows):
        val = str(row[0] or "").strip()
        if "Primer Cuatrimestre" in val:
            if secciones and secciones[-1][2] is None:
                secciones[-1] = (secciones[-1][0], secciones[-1][1], idx)
            secciones.append((1, idx, None))
        elif "Segundo Cuatrimestre" in val:
            if secciones and secciones[-1][2] is None:
                secciones[-1] = (secciones[-1][0], secciones[-1][1], idx)
            secciones.append((2, idx, None))

    if secciones and secciones[-1][2] is None:
        secciones[-1] = (secciones[-1][0], secciones[-1][1], len(all_rows))

    records: list[dict] = []
    for cuatrimestre, start, end in secciones:
        sec_rows = all_rows[start:end]
        # Buscar encabezado "Horario" dentro de la sección
        header_idx = next(
            (j for j, r in enumerate(sec_rows) if str(r[0] or "").strip() == "Horario"),
            None,
        )
        if header_idx is None:
            continue

        data_rows = sec_rows[header_idx + 1:]
        slots = _parse_seccion(data_rows)
        blocks = _merge_blocks(slots)

        for block in blocks:
            records.append({
                "comision_nombre": comision_nombre,
                "cuatrimestre": cuatrimestre,
                "nombre_parcial": block["nombre_parcial"],
                "nombre_3part": block.get("nombre_3part", block["nombre_parcial"]),
                "dia": block["dia"],
                "hora_inicio": block["hora_inicio"],
                "hora_fin": block["hora_fin"],
                "docente_raw": block["docente_raw"],
            })

    return records


# ---------------------------------------------------------------------------
# Seed principal
# ---------------------------------------------------------------------------

def seed(excel_dir: Path, anio: int = 2025, reset: bool = False) -> dict:
    stats: dict = {
        "archivos_procesados": 0,
        "sheets_procesadas": 0,
        "comisiones_creadas": 0,
        "cursadas_creadas": 0,
        "horarios_creados": 0,
        "nombres_no_mapeados": set(),
    }

    with SessionLocal() as db:
        if reset:
            db.execute(delete(Horario))
            db.execute(delete(Cursada))
            db.execute(delete(Comision))
            db.flush()
            print("  [reset] Horarios, cursadas y comisiones borrados.")

        for fname in ARCHIVOS:
            fpath = excel_dir / fname
            if not fpath.exists():
                print(f"  [skip] No encontrado: {fname}")
                continue

            wb = openpyxl.load_workbook(str(fpath))
            stats["archivos_procesados"] += 1

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                records = parse_sheet(ws)
                stats["sheets_procesadas"] += 1

                if not records:
                    continue

                comision_nombre = records[0]["comision_nombre"]

                # Upsert Comision real (ej: "1K01")
                comision = (
                    db.query(Comision)
                    .filter_by(nombre=comision_nombre, anio=anio)
                    .first()
                )
                if comision is None:
                    comision = Comision(nombre=comision_nombre, anio=anio)
                    db.add(comision)
                    db.flush()
                    stats["comisiones_creadas"] += 1

                # Agrupar por (nombre_parcial, cuatrimestre) → una Cursada por materia × cuatrimestre
                grupos: dict[tuple, list[dict]] = defaultdict(list)
                for rec in records:
                    grupos[(rec["nombre_parcial"], rec["cuatrimestre"])].append(rec)

                for (nombre_parcial, cuatrimestre), recs in grupos.items():
                    codigo = mapear_codigo(nombre_parcial)
                    docente_efectivo = next((r["docente_raw"] for r in recs if r["docente_raw"]), None)

                    if codigo is None:
                        # Nombre de 3 partes (p.ej. "Ingeniería y Sociedad")
                        nombre_3p = recs[0].get("nombre_3part", nombre_parcial)
                        codigo = mapear_codigo(nombre_3p)
                        if codigo is not None:
                            docente_efectivo = None

                    if codigo is None:
                        stats["nombres_no_mapeados"].add(nombre_parcial)
                        continue
                    if codigo == "__ignorar__":
                        continue

                    # Upsert Cursada
                    cursada = (
                        db.query(Cursada)
                        .filter_by(
                            comision_id=comision.id,
                            materia_codigo=codigo,
                            cuatrimestre=cuatrimestre,
                        )
                        .first()
                    )
                    if cursada is None:
                        cursada = Cursada(
                            comision_id=comision.id,
                            materia_codigo=codigo,
                            cuatrimestre=cuatrimestre,
                            docente=docente_efectivo,
                        )
                        db.add(cursada)
                        db.flush()
                        stats["cursadas_creadas"] += 1

                    # Upsert Horarios
                    for rec in recs:
                        existe = (
                            db.query(Horario)
                            .filter_by(
                                cursada_id=cursada.id,
                                dia=rec["dia"],
                                hora_inicio=rec["hora_inicio"],
                            )
                            .first()
                        )
                        if existe is None:
                            db.add(
                                Horario(
                                    cursada_id=cursada.id,
                                    dia=rec["dia"],
                                    hora_inicio=rec["hora_inicio"],
                                    hora_fin=rec["hora_fin"],
                                )
                            )
                            stats["horarios_creados"] += 1

            db.commit()

    stats["nombres_no_mapeados"] = sorted(stats["nombres_no_mapeados"])
    return stats


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Seed de comisiones y horarios desde Excel FRRO.")
    parser.add_argument(
        "excel_dir",
        nargs="?",
        default="C:/Users/Bruno/Downloads",
        help="Directorio con los archivos Excel (default: C:/Users/Bruno/Downloads)",
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Borra todas las comisiones y horarios antes de insertar.",
    )
    parser.add_argument(
        "--anio",
        type=int,
        default=2025,
        help="Año académico a asignar a las comisiones (default: 2025).",
    )
    args = parser.parse_args(argv)

    excel_dir = Path(args.excel_dir)
    if not excel_dir.exists():
        print(f"ERROR: Directorio no encontrado: {excel_dir}", file=sys.stderr)
        return 1

    print(f"Leyendo Excel desde: {excel_dir}")
    print(f"Año: {args.anio} | Reset: {args.reset}")
    print()

    stats = seed(excel_dir, anio=args.anio, reset=args.reset)

    print(f"Archivos procesados : {stats['archivos_procesados']}")
    print(f"Sheets procesadas   : {stats['sheets_procesadas']}")
    print(f"Comisiones creadas  : {stats['comisiones_creadas']}")
    print(f"Cursadas creadas    : {stats['cursadas_creadas']}")
    print(f"Horarios creados    : {stats['horarios_creados']}")

    if stats["nombres_no_mapeados"]:
        print(f"\nNombres sin mapear ({len(stats['nombres_no_mapeados'])}):")
        for nombre in stats["nombres_no_mapeados"]:
            print(f"  - {nombre!r}")
    else:
        print("\nTodos los nombres mapeados correctamente.")

    return 0


if __name__ == "__main__":
    sys.exit(_main(sys.argv[1:]))
